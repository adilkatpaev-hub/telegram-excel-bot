import os
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    ConversationHandler,
    ContextTypes,
    MessageHandler,
    filters,
)


LAST_NAME, REPORT_DATE, DEVICE_OR_ADDRESS, WORK_DONE = range(4)

REPORT_FILE = Path("report.xlsx")
HEADERS = [
    "Фамилия",
    "Дата",
    "Наименование устройства или адреса",
    "Проделанная работа",
]


def ensure_report_file() -> None:
    if REPORT_FILE.exists():
        return

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Отчет"
    worksheet.append(HEADERS)
    workbook.save(REPORT_FILE)


def append_report_row(last_name: str, report_date: str, device_or_address: str, work_done: str) -> None:
    ensure_report_file()
    workbook = load_workbook(REPORT_FILE)
    worksheet = workbook.active
    worksheet.append([last_name, report_date, device_or_address, work_done])
    workbook.save(REPORT_FILE)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text("Введите фамилию:")
    return LAST_NAME


async def receive_last_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["last_name"] = update.message.text.strip()
    today = datetime.now().strftime("%d.%m.%Y")
    await update.message.reply_text(f"Введите дату отчета. Например: {today}")
    return REPORT_DATE


async def receive_report_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["report_date"] = update.message.text.strip()
    await update.message.reply_text("Введите наименование устройства или адрес:")
    return DEVICE_OR_ADDRESS


async def receive_device_or_address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["device_or_address"] = update.message.text.strip()
    await update.message.reply_text("Введите проделанную работу:")
    return WORK_DONE


async def receive_work_done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["work_done"] = update.message.text.strip()

    append_report_row(
        context.user_data["last_name"],
        context.user_data["report_date"],
        context.user_data["device_or_address"],
        context.user_data["work_done"],
    )

    await update.message.reply_text(
        "Готово, запись добавлена в report.xlsx.\n"
        "Чтобы добавить новый отчет, отправьте /start."
    )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text("Заполнение отменено. Чтобы начать заново, отправьте /start.")
    return ConversationHandler.END


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "Команды:\n"
        "/start - добавить отчет\n"
        "/cancel - отменить заполнение"
    )


def main() -> None:
    load_dotenv()
    token = os.getenv("BOT_TOKEN")
    if not token:
        raise RuntimeError("Добавьте BOT_TOKEN в файл .env")

    ensure_report_file()

    application = Application.builder().token(token).build()
    conversation = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            LAST_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_last_name)],
            REPORT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_report_date)],
            DEVICE_OR_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_device_or_address)],
            WORK_DONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_work_done)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    application.add_handler(conversation)
    application.add_handler(CommandHandler("help", help_command))
    application.run_polling()


if __name__ == "__main__":
    main()
